import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_inventory = {}
low_stock = {}
supplier_count = 0

for product_row in range(2, product_list.max_row + 1):
    supplier = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_total = product_list.cell(product_row, 5)

    # each supplier product count
    if supplier in product_per_supplier:
        product_per_supplier[supplier] = product_per_supplier[supplier] + 1
    else:
        supplier_count = supplier_count + 1
        product_per_supplier[supplier] = 1

    # total inventory of product per supplier
    if supplier in total_inventory:
        total_inventory[supplier] = total_inventory[supplier] + (inventory * price)
    else:
        total_inventory[supplier] = inventory * price

    # product list that has less than 10 inventory
    if inventory < 10:
        low_stock[int(product_number)] = int(inventory)

    inventory_total.value = inventory * price

print(product_per_supplier)
print(f"Total suppliers are {supplier_count}")
print(total_inventory)
print(low_stock)

inv_file.save("inventory_updated.xlsx")
